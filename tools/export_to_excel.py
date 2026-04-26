"""
Reads tools/descriptions_staging.json and exports a formatted Excel report
to audit_reports/product_descriptions_<timestamp>.xlsx

Usage: py tools/export_to_excel.py
"""
import sys
import json
import os
import re
from datetime import datetime

sys.path.insert(0, ".")

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

STAGING_FILE = "tools/descriptions_staging.json"
OUTPUT_DIR   = "audit_reports"

# Colour palette
GOLD    = "C9A84C"
DARK    = "0D0D0D"
LIGHT   = "F5F0E8"
GREEN   = "3A9E5F"
RED     = "D94040"
AMBER   = "E8943A"
WHITE   = "FFFFFF"
MGRAY   = "AAAAAA"


def strip_html(text):
    return re.sub(r"<[^>]+>", " ", text or "").strip()


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def bold_font(size=10, color=DARK):
    return Font(bold=True, size=size, color=color)


def normal_font(size=9, color=DARK):
    return Font(size=size, color=color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


def thin_border():
    s = Side(style="thin", color="E0D8CC")
    return Border(left=s, right=s, top=s, bottom=s)


def status_color(status):
    if status == "success":   return GREEN
    if status == "failed":    return RED
    if status == "skipped":   return MGRAY
    return AMBER


def main():
    if not os.path.exists(STAGING_FILE):
        print(f"No staging file found at {STAGING_FILE}")
        print("Run /product-descriptions first to generate research data.")
        sys.exit(1)

    with open(STAGING_FILE, encoding="utf-8") as f:
        products = json.load(f)

    if not products:
        print("Staging file is empty.")
        sys.exit(1)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    # Title banner
    ws_sum.merge_cells("A1:I1")
    ws_sum["A1"] = "SCENTIFIED — PRODUCT DESCRIPTION RESEARCH REPORT"
    ws_sum["A1"].font      = Font(bold=True, size=16, color=GOLD)
    ws_sum["A1"].fill      = fill(DARK)
    ws_sum["A1"].alignment = center()
    ws_sum.row_dimensions[1].height = 36

    qa_passed = sum(1 for p in products if p.get("qa_passed"))
    qa_failed = sum(1 for p in products if not p.get("qa_passed"))
    ws_sum.merge_cells("A2:I2")
    ws_sum["A2"] = (
        f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}  |  "
        f"{len(products)} products  |  QA Passed: {qa_passed}  |  QA Failed: {qa_failed}"
    )
    ws_sum["A2"].font      = normal_font(9, MGRAY)
    ws_sum["A2"].fill      = fill("1A1A1A")
    ws_sum["A2"].alignment = center()

    # Header row
    headers    = ["#", "Product Name", "Brand / Category", "Price (USD)",
                  "Notes Found?", "QA", "QA Notes", "Short (chars)", "Long (words)"]
    col_widths = [4, 42, 22, 12, 14, 10, 45, 14, 14]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws_sum.cell(row=4, column=col, value=h)
        cell.font      = bold_font(9, WHITE)
        cell.fill      = fill(DARK)
        cell.alignment = center()
        cell.border    = thin_border()
        ws_sum.column_dimensions[get_column_letter(col)].width = w

    ws_sum.row_dimensions[4].height = 22

    # Data rows
    for i, p in enumerate(products, start=1):
        row        = 4 + i
        bg         = LIGHT if i % 2 == 0 else WHITE
        res        = p.get("research", {})
        notes      = res.get("top_notes") or res.get("heart_notes") or res.get("base_notes")
        short      = strip_html(p.get("short_description", ""))
        long_      = strip_html(p.get("description", ""))
        qa_ok      = p.get("qa_passed", False)
        qa_notes   = p.get("qa_notes", "")
        cats       = ", ".join(p.get("categories", []))

        values = [
            i,
            p.get("name", ""),
            cats,
            p.get("price_usd", ""),
            "Yes" if notes else "No",
            "PASS" if qa_ok else "FAIL",
            qa_notes,
            len(short),
            len(long_.split()),
        ]

        for col, val in enumerate(values, start=1):
            cell = ws_sum.cell(row=row, column=col, value=val)
            cell.fill      = fill(bg)
            cell.border    = thin_border()
            cell.alignment = center() if col in (1, 4, 5, 6, 8, 9) else left()
            cell.font      = normal_font()

        # Colour QA cell
        qa_cell      = ws_sum.cell(row=row, column=6)
        qa_cell.font = bold_font(9, GREEN if qa_ok else RED)

        ws_sum.row_dimensions[row].height = 28

    # ── Sheet 2: Research Data ────────────────────────────────────────────────
    ws_res = wb.create_sheet("Research Data")
    ws_res.sheet_view.showGridLines = False

    ws_res.merge_cells("A1:G1")
    ws_res["A1"] = "FRAGRANCE RESEARCH — NOTES & SOURCES"
    ws_res["A1"].font      = Font(bold=True, size=14, color=GOLD)
    ws_res["A1"].fill      = fill(DARK)
    ws_res["A1"].alignment = center()
    ws_res.row_dimensions[1].height = 30

    res_headers  = ["Product Name", "Top Notes", "Heart Notes", "Base Notes",
                    "Longevity", "Projection", "Sources Used"]
    res_widths   = [42, 28, 28, 28, 16, 16, 50]

    for col, (h, w) in enumerate(zip(res_headers, res_widths), start=1):
        cell = ws_res.cell(row=3, column=col, value=h)
        cell.font      = bold_font(9, WHITE)
        cell.fill      = fill(DARK)
        cell.alignment = center()
        cell.border    = thin_border()
        ws_res.column_dimensions[get_column_letter(col)].width = w

    ws_res.row_dimensions[3].height = 22

    for i, p in enumerate(products, start=1):
        row = 3 + i
        bg  = LIGHT if i % 2 == 0 else WHITE
        res = p.get("research", {})
        sources = ", ".join(res.get("sources_used", []))

        values = [
            p.get("name", ""),
            res.get("top_notes", "Not found"),
            res.get("heart_notes", "Not found"),
            res.get("base_notes", "Not found"),
            res.get("longevity", "Unknown"),
            res.get("projection", "Unknown"),
            sources or "No external sources found",
        ]

        for col, val in enumerate(values, start=1):
            cell = ws_res.cell(row=row, column=col, value=val)
            cell.fill      = fill(bg)
            cell.border    = thin_border()
            cell.alignment = left()
            cell.font      = normal_font()

        ws_res.row_dimensions[row].height = 32

    # ── Sheet 3: Generated Copy ───────────────────────────────────────────────
    ws_copy = wb.create_sheet("Generated Copy")
    ws_copy.sheet_view.showGridLines = False

    ws_copy.merge_cells("A1:D1")
    ws_copy["A1"] = "GENERATED SEO DESCRIPTIONS — REVIEW BEFORE GOING LIVE"
    ws_copy["A1"].font      = Font(bold=True, size=14, color=GOLD)
    ws_copy["A1"].fill      = fill(DARK)
    ws_copy["A1"].alignment = center()
    ws_copy.row_dimensions[1].height = 30

    copy_headers = ["Product Name", "Short Description", "Long Description (plain text)", "Update Status"]
    copy_widths  = [38, 55, 80, 14]

    for col, (h, w) in enumerate(zip(copy_headers, copy_widths), start=1):
        cell = ws_copy.cell(row=3, column=col, value=h)
        cell.font      = bold_font(9, WHITE)
        cell.fill      = fill(DARK)
        cell.alignment = center()
        cell.border    = thin_border()
        ws_copy.column_dimensions[get_column_letter(col)].width = w

    ws_copy.row_dimensions[3].height = 22

    for i, p in enumerate(products, start=1):
        row    = 3 + i
        bg     = LIGHT if i % 2 == 0 else WHITE
        short  = strip_html(p.get("short_description", ""))
        long_  = strip_html(p.get("description", ""))
        status = p.get("update_status", "pending")

        values = [p.get("name", ""), short, long_, status.upper()]

        for col, val in enumerate(values, start=1):
            cell = ws_copy.cell(row=row, column=col, value=val)
            cell.fill      = fill(bg)
            cell.border    = thin_border()
            cell.alignment = left()
            cell.font      = normal_font()

        status_cell      = ws_copy.cell(row=row, column=4)
        status_cell.font = bold_font(9, status_color(status))
        status_cell.alignment = center()
        ws_copy.row_dimensions[row].height = 80

    # ── Sheet 4: Raw HTML ─────────────────────────────────────────────────────
    ws_html = wb.create_sheet("Raw HTML")
    ws_html.sheet_view.showGridLines = False

    ws_html.merge_cells("A1:C1")
    ws_html["A1"] = "RAW HTML — COPY THIS DIRECTLY INTO WOOCOMMERCE IF NEEDED"
    ws_html["A1"].font      = Font(bold=True, size=14, color=GOLD)
    ws_html["A1"].fill      = fill(DARK)
    ws_html["A1"].alignment = center()
    ws_html.row_dimensions[1].height = 30

    html_headers = ["Product Name", "Short Description (plain)", "Long Description (HTML)"]
    html_widths  = [38, 55, 100]

    for col, (h, w) in enumerate(zip(html_headers, html_widths), start=1):
        cell = ws_html.cell(row=3, column=col, value=h)
        cell.font      = bold_font(9, WHITE)
        cell.fill      = fill(DARK)
        cell.alignment = center()
        cell.border    = thin_border()
        ws_html.column_dimensions[get_column_letter(col)].width = w

    ws_html.row_dimensions[3].height = 22

    for i, p in enumerate(products, start=1):
        row   = 3 + i
        bg    = LIGHT if i % 2 == 0 else WHITE
        values = [
            p.get("name", ""),
            p.get("short_description", ""),
            p.get("description", ""),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws_html.cell(row=row, column=col, value=val)
            cell.fill      = fill(bg)
            cell.border    = thin_border()
            cell.alignment = left()
            cell.font      = normal_font()
        ws_html.row_dimensions[row].height = 80

    # ── Save ──────────────────────────────────────────────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = os.path.join(OUTPUT_DIR, f"product_descriptions_{ts}.xlsx")
    wb.save(out_path)
    print(f"Excel report saved: {out_path}")
    return out_path


if __name__ == "__main__":
    main()
