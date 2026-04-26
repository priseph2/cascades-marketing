"""
Exports seo_staging.json to a 3-sheet Excel report.

Usage: py tools/export_seo_excel.py
"""
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

STAGING_FILE = "tools/seo_staging.json"
OUTPUT_DIR   = "audit_reports"

GREEN  = "C6EFCE"
RED    = "FFC7CE"
YELLOW = "FFEB9C"
HEADER = "1F4E79"
ALT    = "EBF3FB"


def col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def header_row(ws, cols, row=1):
    fill   = PatternFill("solid", fgColor=HEADER)
    font   = Font(bold=True, color="FFFFFF")
    border = Border(bottom=Side(style="medium"))
    for i, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill   = fill
        c.font   = font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main():
    with open(STAGING_FILE, "r", encoding="utf-8") as f:
        staging = json.load(f)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.row_dimensions[1].height = 30

    cols1 = ["ID", "Product Name", "Brand", "Price (₦)",
             "Focus Keyword",
             "Title chars", "Desc chars",
             "Slug (current)", "Slug (proposed)",
             "Image Alt Text",
             "QA", "QA Notes"]
    header_row(ws1, cols1)

    for i, e in enumerate(staging, 2):
        brand = e.get("categories", [""])[0] if e.get("categories") else ""
        pro_title = e.get("proposed_seo_title", "")
        pro_desc  = e.get("proposed_seo_desc", "")
        qa = "PASS" if e.get("qa_passed") else "FAIL"
        fill_color = GREEN if qa == "PASS" else RED
        alt_fill   = PatternFill("solid", fgColor=ALT if i % 2 == 0 else "FFFFFF")
        qa_fill    = PatternFill("solid", fgColor=fill_color)

        row = [
            e.get("id"), e.get("name"), brand, e.get("price", ""),
            e.get("focus_keyword", ""),
            len(pro_title), len(pro_desc),
            e.get("current_slug", ""), e.get("proposed_slug", ""),
            e.get("image_alt_text", ""),
            qa, e.get("qa_notes", ""),
        ]
        for j, val in enumerate(row, 1):
            c = ws1.cell(row=i, column=j, value=val)
            c.alignment = Alignment(vertical="center", wrap_text=(j in (2, 12)))
            c.fill = qa_fill if j == 11 else alt_fill

    for col, w in zip(range(1, len(cols1)+1), [8, 38, 18, 12, 28, 10, 10, 35, 35, 30, 8, 40]):
        col_width(ws1, col, w)
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Side by Side ───────────────────────────────────────────
    ws2 = wb.create_sheet("Side by Side")
    ws2.row_dimensions[1].height = 30

    cols2 = ["ID", "Product Name", "Brand", "Focus Keyword",
             "Current SEO Title", "Proposed SEO Title",
             "Current Meta Description", "Proposed Meta Description",
             "Image Alt Text", "Description Patched?"]
    header_row(ws2, cols2)

    for i, e in enumerate(staging, 2):
        brand = e.get("categories", [""])[0] if e.get("categories") else ""
        alt_fill = PatternFill("solid", fgColor=ALT if i % 2 == 0 else "FFFFFF")
        desc_patched = "Yes" if e.get("proposed_description") else "No"
        row = [
            e.get("id"), e.get("name"), brand,
            e.get("focus_keyword", ""),
            e.get("current_seo_title", ""), e.get("proposed_seo_title", ""),
            e.get("current_seo_desc", ""),  e.get("proposed_seo_desc", ""),
            e.get("image_alt_text", ""),
            desc_patched,
        ]
        for j, val in enumerate(row, 1):
            c = ws2.cell(row=i, column=j, value=val)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.fill = alt_fill
            if j in (5, 6):
                c.fill = PatternFill("solid", fgColor=("FFF2CC" if j == 5 else "E2EFDA") if i % 2 == 0 else ("FFFBEA" if j == 5 else "F0F7EC"))

    for col, w in zip(range(1, len(cols2)+1), [8, 33, 16, 25, 33, 33, 50, 50, 28, 14]):
        col_width(ws2, col, w)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Research Data ──────────────────────────────────────────
    ws3 = wb.create_sheet("Research Data")
    ws3.row_dimensions[1].height = 30

    cols3 = ["ID", "Product Name", "Primary Keyword", "Sources Used",
             "Keyword Search Results", "Proposed SEO Title", "Proposed Meta Desc",
             "QA Notes", "Push Status"]
    header_row(ws3, cols3)

    for i, e in enumerate(staging, 2):
        alt_fill = PatternFill("solid", fgColor=ALT if i % 2 == 0 else "FFFFFF")
        research = e.get("research", {})
        row = [
            e.get("id"), e.get("name"),
            research.get("primary_keyword", ""),
            "\n".join(research.get("sources_used", [])),
            research.get("serp_notes", ""),
            e.get("proposed_seo_title", ""),
            e.get("proposed_seo_desc", ""),
            e.get("qa_notes", ""),
            e.get("update_status", "draft"),
        ]
        for j, val in enumerate(row, 1):
            c = ws3.cell(row=i, column=j, value=val)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.fill = alt_fill

    for col, w in zip(range(1, len(cols3)+1), [8, 35, 25, 40, 40, 35, 55, 40, 12]):
        col_width(ws3, col, w)
    ws3.freeze_panes = "A2"

    # ── Save ────────────────────────────────────────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(OUTPUT_DIR, f"seo_gaps_{ts}.xlsx")
    wb.save(path)
    print(path)


if __name__ == "__main__":
    main()
